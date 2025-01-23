import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import os
from pathlib import Path

def classify_fate(notes):
    try:
        notes = notes.replace("\n", ".")
        if isinstance(notes, str):
            # Rules based on notes to determine fate
            if any(x in notes.lower() for x in ["combat loss", "shot down", "lost", "loss", "lost in action"]):
                return "X"
            elif any(x in notes.lower() for x in
                     ["crashed", "wrecked", "crash landing", "hard landing", "condemned", "crash",
                      "written off after crash landing", "ditched", "takeoff accident", "midair collision", "surplused",
                      "damaged landing", "damaged in landing", "surveyed", "damaged beyond repair", "damaged beyond repair"]):
                return "C"
            elif any(x in notes.lower() for x in
                     ["decommissioned", "retired", "written off", "w/o", "withdrawn", "wfu", "dismantled",
                      "struck off charge", "soc", "scrapped", "off inventory", "salvaged", "MASDC", "reclamation"]):
                return "D"
            if any(x in notes.lower() for x in
                   ["sold", "private ownership", "private ownership", "lend-leased", "acquired", "civilian",
                    "private transaction", "as a sale", "donated", "CAA", "RFC"]):
                return "S"
            elif any(x in notes.lower() for x in ["cl-26", "trainer", "target", "instruction", "instructional"]):
                return "Q"
            elif any(x in notes.lower() for x in
                     ["transferred", "delivered", "diverted", "raf", "uk", "canada", "rcaf", "philippines", "netherlands", "france", "ussr",
                      "rfc"]):
                return "T"
            else:
                return "O"
    except AttributeError:
        print("Unable to classify fate.", notes)
        return 'Z'


def format_column_fate(filepath):
    # Charger le classeur et la feuille active
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    # Créer un style de remplissage rouge/vert pour la colonne F
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
    column = 'F'

    # Parcourir les cellules de la colonne 'G'
    for row in range(2, sheet.max_row + 1):  # en commençant par 2 si votre feuille a des en-têtes
        cell = sheet[f'{column}{row}']
        if cell.value == False:
            cell.fill = red_fill
        if cell.value == True:
            cell.fill = green_fill

    # Sauvegarder les modifications dans un nouveau fichier
    workbook.save(filepath)


def nice_print_format(str):
    n = len(str)
    line = "-" * n
    print(f"{line}\n{str}\n{line}")


def recheck_fate(data, file):
    data['Calculated Fate'] = data['Notes'].apply(classify_fate)
    data['Unchanged Fate'] = data['Calculated Fate'].str.strip() == data['Fate'].str.strip()
    data[['Serial Number', 'Build', 'Type', 'Fate', 'Calculated Fate', 'Unchanged Fate', 'Notes']].to_excel(
        f"./output/{file}_fate.xlsx", index=False)

    nice_print_format(data['Unchanged Fate'].value_counts())
    format_column_fate(f"./output/{file}_fate.xlsx")


if __name__ == "__main__":
    while True:
        ## Récupérer les noms de fichiers disponibles dans /input
        files = [f for f in os.listdir('./extracted_data') if f.endswith('.xlsx') and 'fate' not in f]

        ## Trier ces fichiers par ordre ascendant
        files.sort()

        ## Afficher un menu qui va permettre de choisir avec quel fichier on veut travailler
        nice_print_format("Choose a file to work with :")
        for i, file in enumerate(files):
            print(f"{i + 1}. {file}")
        choice = int(input("Enter the number of the file you want to work with : "))
        if choice < 1 or choice > len(files):
            nice_print_format("\nInvalid choice. Please try again.")
            continue

        filename = f"./output/{files[choice - 1]}"
        name = Path(filename).stem
        try:
            nice_print_format(f"\nChecking fate and formatting for  {filename}...\n")

            sorted_data = pd.read_excel(f"{filename}")
            recheck_fate(sorted_data, name)
            nice_print_format("\nSecond check of fate outputs completed.\n")
        except FileNotFoundError as e:
            nice_print_format("\nFile not found. Please try again.\n", e)
            continue